import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import config
from typing import List, Dict, Optional, Union
from io import BytesIO

class ExcelHandler:
    @staticmethod
    def create_input_template(file_path: Optional[Union[str, BytesIO]] = None):
        """
        创建输入Excel模板
        
        Args:
            file_path: 模板文件保存路径，可以是字符串路径或BytesIO对象
        """
        if file_path is None:
            file_path = config.OUTPUT_EXCEL_TEMPLATE
            
        # 创建示例数据
        sample_data = {
            "元件型号": ["LM358DR", "ESP32-WROOM-32D", "TL072CDR", "", ""],
            "元件描述": ["运算放大器", "WiFi蓝牙模块", "低噪声JFET双运放", "", ""]
        }
        
        df = pd.DataFrame(sample_data)
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "元件查询"
        
            # 添加数据
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")
        
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
            # 格式化标题行
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border
            
            # 格式化数据行
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = alignment
                    cell.border = thin_border
            
            # 自动调整列宽
            for i, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(i)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            if isinstance(file_path, BytesIO):
                wb.save(file_path)
            else:
                wb.save(file_path)
                print(f"输入模板已创建: {file_path}")
    
    @staticmethod
    def create_result_template(results: List[Dict], file_path: Optional[Union[str, BytesIO]] = None):
        """
        创建结果Excel文件
        
        Args:
            results: 查询结果列表
            file_path: 结果文件保存路径，可以是字符串路径或BytesIO对象
        """
        if file_path is None:
            file_path = config.OUTPUT_EXCEL_RESULT
            
        # 定义正确的列顺序（与GUI中显示的顺序完全一致）
        column_order = [
            "元件型号", "搜索型号", "产品名称", "品牌", 
            "价格", "最大批次", "库存", "是否停产", 
            "替代型号", "备注"
        ]
        
        # 定义中文列名映射
        chinese_columns = {
            "元件型号": "元件型号",
            "搜索型号": "搜索型号", 
            "产品名称": "产品名称",
            "品牌": "品牌",
            "价格": "价格(CNY)",
            "最大批次": "最大批次(pcs)",
            "库存": "库存",
            "是否停产": "是否停产",
            "替代型号": "替代型号",
            "备注": "备注"
        }
        
        # 创建DataFrame
        df = pd.DataFrame(results)
        
        # 确保列顺序正确（关键修复点）
        # 先选择需要的列，再重新排序
        df_selected = df[column_order]
        
        # 重命名列
        df_final = df_selected.rename(columns=chinese_columns)  # type: ignore
        
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = "查询结果"
        
            # 添加数据
            for r in dataframe_to_rows(df_final, index=False, header=True):
                ws.append(r)
            
            # 设置样式
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center")
        
            # 设置边框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
            # 格式化标题行
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = alignment
                cell.border = thin_border
            
            # 格式化数据行
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = thin_border
                
                    # 对价格列进行特殊格式化（第5列，openpyxl列索引从1开始）
                    if cell.column == 5:  # 价格列（价格(CNY)）
                        if isinstance(cell.value, (int, float)) and cell.value > 0:
                            cell.number_format = '"¥"#,##0.00000'
                        elif isinstance(cell.value, (int, float)):
                            cell.number_format = '"¥"0'
                
                    # 对批次数量列进行特殊格式化（第6列，openpyxl列索引从1开始）
                    if cell.column == 6:  # 批次数量列（最大批次(pcs)）
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
            
            # 自动调整列宽
            for i, column in enumerate(ws.columns, 1):
                max_length = 0
                column_letter = get_column_letter(i)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 保存文件
            if isinstance(file_path, BytesIO):
                wb.save(file_path)
            else:
                wb.save(file_path)
                print(f"结果文件已创建: {file_path}")
    
    @staticmethod
    def read_components_from_excel(file_path: str) -> List[str]:
        """
        从Excel文件读取电子元器件型号
        
        Args:
            file_path: Excel文件路径
            
        Returns:
            电子元器件型号列表
        """
        try:
            df = pd.read_excel(file_path, sheet_name=0)  # 读取第一个工作表
            # 尝试不同的列名
            possible_columns = ["元件型号", "Part Number", "型号", "元件编号"]
            part_numbers = []
            
            for col in possible_columns:
                if col in df.columns:
                    part_numbers = df[col].dropna().tolist()
                    break
            
            if not part_numbers:
                # 如果没有找到标准列名，使用第一列
                first_column = df.columns[0]
                part_numbers = df[first_column].dropna().tolist()
                
            return [str(part_num).strip() for part_num in part_numbers if str(part_num).strip()]
        except Exception as e:
            print(f"读取Excel文件时发生错误: {str(e)}")
            return []
    
    @staticmethod
    def read_components_from_txt(file_path: str) -> List[str]:
        """
        从txt文件读取电子元器件型号
        
        Args:
            file_path: txt文件路径
            
        Returns:
            电子元器件型号列表
        """
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                lines = f.readlines()
            # 去除空白行和首尾空格
            return [line.strip() for line in lines if line.strip()]
        except Exception as e:
            print(f"读取txt文件时发生错误: {str(e)}")
            return []